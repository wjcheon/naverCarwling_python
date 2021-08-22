from selenium import webdriver
from webdriver_manager.firefox import GeckoDriverManager
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time

import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook
from random import *
# tokenizer
from konlpy.tag import Okt  # Tokenizer
okt = Okt()
# data import
import pandas as pd
import numpy as np
import math

## PARAMETERS
endSearchingPage = 5
# 크롤링 시작 일자: 날짜는 반드시 0000.00.00 자릿수를 맞춰줘야 한다! (2021.08.18)
stDate = '2021.07.01'
# 크롤링 종료 일자
endDate = '2021.07.31'

dbPath = "/Users/wjcheon/Dropbox/WeKnew/naver_kin_crawling-master/질병목록.xlsx"
df = pd.read_excel(dbPath, engine='openpyxl')
diseaseList = df.iloc[:, 1]
diseaseList = np.array(diseaseList)

# firefox 버전
profile = webdriver.FirefoxProfile()
profile.set_preference('general.useragent.override', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:65.0) Gecko/20100101 Firefox/65.0')
profile.set_preference("network.proxy.type", 1)
profile.set_preference("network.proxy.socks", "127.0.0.1")
profile.set_preference("network.proxy.socks_port", 9050)   # Proxy set as 9050 port

#path = "/Users/taehyung/anaconda3/envs/study/geckodriver"
#path = '/Users/wjcheon/PycharmProjects/naver_kin_crawling/driver'
#driver = webdriver.Firefox(firefox_profile=profile, executable_path=path)
driver = webdriver.Firefox(firefox_profile=profile, executable_path=GeckoDriverManager().install())  # automatic install path was used

# 네이버 지식인 크롤링
# keyword에 크롤링하고 싶은 단어 선택. space 는 + 로 치환
def get_keyword(text):
    return text.replace(" ", "%20")

# 정렬 방식 선택
# 1: 추천순
# 2: 최신순
# 기타: 정확도 순
def sort_kind(index):
    # 추천
    if index == 1:
        return 'vcount'
    # 최신순
    elif index == 2:
        return 'date'
    # 정확도
    else:
        return 'none'
counter =0
for iterKeyword in diseaseList:
    counter += 1
    print("{} is under the DB generating process | {}/{}".format(iterKeyword, counter, len(diseaseList)))
    if not pd.isna(iterKeyword):
        keyword = iterKeyword
        driver.get('https://kin.naver.com/search/list.nhn?query=' + get_keyword(keyword))
        time.sleep(uniform(0.1, 1.0)) # trick for blocking

        page_index = 1 # 1 is initial value
        period_txt = "&period=" + stDate + ".%7C" + endDate + "."

        _sort_kind = sort_kind(3) # 3 is 정확도순
        date = str(datetime.now()).replace('.', '_')
        date = date.replace(' ', '_')

        # URL 저장
        f = open("result/url_list" + "_" + keyword.replace(' ', '+') + "_" + date + ".txt", 'w')
        page_url = []
        counter2 = 0
        while True:
            counter2 += 1
            print("{} th page URL is collected !!".format(counter2))
            time.sleep(uniform(0.01, 1.0))
            # Correct:https://kin.naver.com/search/list.nhn?sort=date&query=%EB%8B%B9%EB%87%A8&period=2021.07.01.%7C2021.07.31.&section=kin&page=2
            # Fail:   https://kin.naver.com/search/list.nhn?sort=date&query=%EB%8B%B9%EB%87%A8&period=2021.7.01.%7C2021.7.31.&section=kin&page=2
            # Fail:   https://kin.naver.com/search/list.nhn?&sort=date&section=kin&query=%EB%8B%B9%EB%87%A8&period=2021.7.01.%7C2021.7.31.&page=2
            # original
            #driver.get('https://kin.naver.com/search/list.nhn?' + "&sort=" + _sort_kind + '&query=' + get_keyword(keyword) + period_txt + "&section=kin" + "&page=" + str(page_index))
            # second
            driver.get('https://kin.naver.com/search/list.nhn?' + "sort=" + _sort_kind + '&query=' + get_keyword(
                keyword) + period_txt + "&section=kin" + "&page=" + str(page_index))
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            tags = soup.find_all('a', class_="_nclicks:kin.txt _searchListTitleAnchor")
            for tag in tags:
                url = str(tag).split(' ')[3]
                url = url.replace('href=', "")
                url = url.replace('"', "")
                url = url.replace('amp;', '')
                page_url.append(url)
                f.write(url + "\n")

            post_number = driver.find_element_by_class_name('number').text
            post_number = str(post_number).replace("(", "")
            post_number = str(post_number).replace(")", "")

            current_number = post_number.split('/')[0].split('-')[1]
            current_number = current_number.replace(',', '')
            total_number = post_number.split('/')[1]
            total_number = total_number.replace(',', '')

            # wjcheon: for stopping trigger
            #if int(current_number) == int(total_number):
            if page_index==5:
                break
            else:
                page_index += 1
        print("The {} numbers of contents were collected !!".format(len(page_url)))
        filename = 'result/' + keyword.replace(' ', '.') + "_" + date + "_crawling_result.xlsx"

        wb = Workbook()
        sheet = wb.active
        sheet.append(['제목', '질문', '답변'])

        for j in range(1, 4):
            sheet.cell(row=1, column=j).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
        count=0
        for i in page_url:

            try:
                driver.get(i)
                title = driver.find_element_by_class_name('title').text  # 제목
                question_txt = driver.find_element_by_class_name('c-heading__content').text # 질문

                title_tokenizer = okt.morphs(title) # title 과 keyword를 이용해서 screening을 함
                # if keyword not in title_tokenizer:
                if not any(keyword in s for s in title_tokenizer):
                    print('pass: by a title')
                    continue

            except:
                question_txt = ""
                continue

            # 답변 리스트
            answer_list = driver.find_elements_by_class_name("se-main-container")
            try:
                for n, answer in enumerate(answer_list):
                    texts = answer.find_elements_by_tag_name('span')
                    t = ""
                    for i in texts:
                        t += i.text

                    if n == 0:
                        sheet.append([title, question_txt, t])
                    else:
                        sheet.append(["", "", t])
            except:
                print('pass: by a answer')
                continue

            count += 1
            print("{} th contents is under analyzing..".format(count))  # print status

        wb.save(filename)
        print("DB:{} was successfully generated !!".format(iterKeyword))
    else:
        continue